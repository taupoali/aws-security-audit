#!/usr/bin/env python3

import csv
import glob
import os
import argparse
from datetime import datetime

def get_csv_files(directory):
    """Get all CSV files in directory"""
    pattern = os.path.join(directory, "*.csv")
    return glob.glob(pattern)

def sanitize_sheet_name(filename):
    """Sanitize filename for Excel sheet name"""
    # Remove extension and path
    name = os.path.splitext(os.path.basename(filename))[0]
    
    # Excel sheet name restrictions
    invalid_chars = ['\\', '/', '*', '[', ']', ':', '?']
    for char in invalid_chars:
        name = name.replace(char, '_')
    
    # Limit to 31 characters (Excel limit)
    if len(name) > 31:
        name = name[:31]
    
    return name

def load_csv_data(file_path):
    """Load CSV data"""
    try:
        with open(file_path, 'r', newline='', encoding='utf-8') as f:
            reader = csv.reader(f)
            return list(reader)
    except Exception as e:
        print(f"Error loading {file_path}: {e}")
        return []

def write_excel_with_openpyxl(csv_files, output_file):
    """Write Excel file using openpyxl"""
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)
        
        for csv_file in csv_files:
            sheet_name = sanitize_sheet_name(csv_file)
            data = load_csv_data(csv_file)
            
            if not data:
                continue
            
            ws = wb.create_sheet(title=sheet_name)
            
            # Write data
            for row_idx, row_data in enumerate(data, 1):
                for col_idx, cell_value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=cell_value)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            print(f"[INFO] Added sheet: {sheet_name}")
        
        wb.save(output_file)
        return True
        
    except ImportError:
        return False

def write_excel_with_xlsxwriter(csv_files, output_file):
    """Write Excel file using xlsxwriter"""
    try:
        import xlsxwriter
        
        workbook = xlsxwriter.Workbook(output_file)
        
        # Create header format
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#D3D3D3',
            'border': 1
        })
        
        for csv_file in csv_files:
            sheet_name = sanitize_sheet_name(csv_file)
            data = load_csv_data(csv_file)
            
            if not data:
                continue
            
            worksheet = workbook.add_worksheet(sheet_name)
            
            # Write data
            for row_idx, row_data in enumerate(data):
                for col_idx, cell_value in enumerate(row_data):
                    if row_idx == 0:  # Header row
                        worksheet.write(row_idx, col_idx, cell_value, header_format)
                    else:
                        worksheet.write(row_idx, col_idx, cell_value)
            
            # Auto-adjust column widths
            if data:
                for col_idx in range(len(data[0])):
                    max_width = 0
                    for row in data:
                        if col_idx < len(row):
                            width = len(str(row[col_idx]))
                            if width > max_width:
                                max_width = width
                    worksheet.set_column(col_idx, col_idx, min(max_width + 2, 50))
            
            print(f"[INFO] Added sheet: {sheet_name}")
        
        workbook.close()
        return True
        
    except ImportError:
        return False

def write_excel_manual(csv_files, output_file):
    """Write Excel file manually using XML format"""
    import xml.etree.ElementTree as ET
    import zipfile
    import tempfile
    import shutil
    
    # Create temporary directory
    temp_dir = tempfile.mkdtemp()
    
    try:
        # Create basic Excel structure
        os.makedirs(os.path.join(temp_dir, "_rels"))
        os.makedirs(os.path.join(temp_dir, "xl", "_rels"))
        os.makedirs(os.path.join(temp_dir, "xl", "worksheets"))
        
        # Create [Content_Types].xml
        content_types = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
<Default Extension="xml" ContentType="application/xml"/>
<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'''
        
        for i in range(len(csv_files)):
            content_types += f'<Override PartName="/xl/worksheets/sheet{i+1}.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        
        content_types += '</Types>'
        
        with open(os.path.join(temp_dir, "[Content_Types].xml"), 'w') as f:
            f.write(content_types)
        
        # Create _rels/.rels
        rels = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
</Relationships>'''
        
        with open(os.path.join(temp_dir, "_rels", ".rels"), 'w') as f:
            f.write(rels)
        
        # Create xl/_rels/workbook.xml.rels
        workbook_rels = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'''
        
        for i in range(len(csv_files)):
            workbook_rels += f'<Relationship Id="rId{i+1}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet{i+1}.xml"/>'
        
        workbook_rels += '</Relationships>'
        
        with open(os.path.join(temp_dir, "xl", "_rels", "workbook.xml.rels"), 'w') as f:
            f.write(workbook_rels)
        
        # Create xl/workbook.xml
        workbook_xml = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
<sheets>'''
        
        for i, csv_file in enumerate(csv_files):
            sheet_name = sanitize_sheet_name(csv_file)
            workbook_xml += f'<sheet name="{sheet_name}" sheetId="{i+1}" r:id="rId{i+1}" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"/>'
        
        workbook_xml += '</sheets></workbook>'
        
        with open(os.path.join(temp_dir, "xl", "workbook.xml"), 'w') as f:
            f.write(workbook_xml)
        
        # Create worksheets
        for i, csv_file in enumerate(csv_files):
            data = load_csv_data(csv_file)
            if not data:
                continue
            
            worksheet_xml = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
<sheetData>'''
            
            for row_idx, row_data in enumerate(data, 1):
                worksheet_xml += f'<row r="{row_idx}">'
                for col_idx, cell_value in enumerate(row_data):
                    col_letter = chr(65 + col_idx) if col_idx < 26 else chr(65 + col_idx // 26 - 1) + chr(65 + col_idx % 26)
                    cell_ref = f"{col_letter}{row_idx}"
                    # Escape XML characters
                    cell_value = str(cell_value).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                    worksheet_xml += f'<c r="{cell_ref}" t="inlineStr"><is><t>{cell_value}</t></is></c>'
                worksheet_xml += '</row>'
            
            worksheet_xml += '</sheetData></worksheet>'
            
            with open(os.path.join(temp_dir, "xl", "worksheets", f"sheet{i+1}.xml"), 'w') as f:
                f.write(worksheet_xml)
            
            sheet_name = sanitize_sheet_name(csv_file)
            print(f"[INFO] Added sheet: {sheet_name}")
        
        # Create ZIP file
        with zipfile.ZipFile(output_file, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for root, dirs, files in os.walk(temp_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    arc_path = os.path.relpath(file_path, temp_dir)
                    zipf.write(file_path, arc_path)
        
        return True
        
    finally:
        # Clean up temp directory
        shutil.rmtree(temp_dir)

def main():
    parser = argparse.ArgumentParser(description="Convert CSV files to Excel workbook with multiple tabs")
    parser.add_argument("--input-dir", default=".", help="Directory containing CSV files")
    parser.add_argument("--output", default="consolidated_data.xlsx", help="Output Excel file")
    args = parser.parse_args()
    
    print(f"[INFO] Scanning for CSV files in: {args.input_dir}")
    
    # Find all CSV files
    csv_files = get_csv_files(args.input_dir)
    
    if not csv_files:
        print("[ERROR] No CSV files found")
        return
    
    print(f"[INFO] Found {len(csv_files)} CSV files")
    
    # Try different Excel writing methods
    success = False
    
    # Method 1: Try openpyxl
    print("[INFO] Attempting to use openpyxl...")
    success = write_excel_with_openpyxl(csv_files, args.output)
    
    if not success:
        # Method 2: Try xlsxwriter
        print("[INFO] openpyxl not available, trying xlsxwriter...")
        success = write_excel_with_xlsxwriter(csv_files, args.output)
    
    if not success:
        # Method 3: Manual XML creation
        print("[INFO] Excel libraries not available, using manual XML method...")
        try:
            success = write_excel_manual(csv_files, args.output)
        except Exception as e:
            print(f"[ERROR] Manual Excel creation failed: {e}")
            success = False
    
    if success:
        print(f"[INFO] Excel file created successfully: {args.output}")
        print(f"[INFO] Total sheets: {len(csv_files)}")
    else:
        print("[ERROR] Failed to create Excel file. Please install openpyxl or xlsxwriter:")
        print("pip install openpyxl")
        print("or")
        print("pip install xlsxwriter")

if __name__ == "__main__":
    main()